import streamlit as st
from streamlit import caching

import pandas as pd
import openpyxl

import seaborn as sns

# streamlit
st.set_page_config(page_title='Portfolio Analysis', page_icon=':moneybag:', layout='wide', initial_sidebar_state='expanded')
st.title(':moneybag: ' + 'Portfolio Analysis')


# Upload File & Parse Data
uploaded_file = st.sidebar.file_uploader('Upload Holdings file from Zerodha here', type=['xlsx']) #accept_multiple_files=True
if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file)
    sheetname = st.sidebar.selectbox('Select Sheets', wb.sheetnames)
    ws = wb['{}'.format(sheetname)]
    ls = []

    cell_obj = ws['B23': 'M{}'.format(ws.max_row)]
    for rows in cell_obj:
        for row in rows:
            ls.append(row.value)

    chunked_list = list()
    chunk_size = 12
    for i in range(0, len(ls), chunk_size):
        chunked_list.append(ls[i:i+chunk_size])
    df = pd.DataFrame(chunked_list)
    
    try:
        df.columns = df.iloc[0]
    except:
        mutual_fund_holdings = 1
        st.warning('No Mutual Fund Holdings')

    # Drop Unrequired Rows and Columns
    df1 = df.drop(columns=['ISIN', 'Quantity Available', 'Sector', 'Quantity Discrepant', 'Quantity Long Term', 'Quantity Pledged (Margin)', 'Quantity Pledged (Loan)'])
    df1 = df1.drop([0])

    df1['Total Quantity'] = df['Quantity Available'] + df['Quantity Discrepant'] + df['Quantity Long Term'] + df['Quantity Pledged (Margin)'] + df['Quantity Pledged (Loan)']
    df1['Invested Value'] = df1['Total Quantity'] * df1['Average Price']
    df1['Current Value']  = df1['Total Quantity'] * df1['Previous Closing Price']
    
    st.dataframe(df1)